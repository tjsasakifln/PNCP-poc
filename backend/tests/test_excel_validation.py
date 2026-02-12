"""Comprehensive Excel validation test suite (STORY-202 TEST-04)."""
import pytest
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

from excel import create_excel, sanitize_for_excel, parse_datetime


def _make_bid(**overrides):
    """Create a sample bid dictionary with defaults."""
    bid = {
        "codigoCompra": "12345678000190-1-000001/2025",
        "objetoCompra": "Aquisição de uniformes para funcionários",
        "nomeOrgao": "Prefeitura Municipal de São Paulo",
        "uf": "SP",
        "municipio": "São Paulo",
        "valorTotalEstimado": 150000.00,
        "modalidadeNome": "Pregão - Eletrônico",
        "dataPublicacaoPncp": "2025-01-15",
        "dataAberturaProposta": "2025-02-01T10:00:00",
        "situacaoCompraNome": "Aberto",
        "numeroControlePNCP": "12345678000190-1-000001/2025",
        "linkSistemaOrigem": "https://example.gov.br/edital/123",
    }
    bid.update(overrides)
    return bid


class TestCreateExcelStructure:
    """Test Excel file structure and formatting."""

    def test_returns_bytesio(self):
        result = create_excel([_make_bid()])
        assert isinstance(result, BytesIO)

    def test_valid_xlsx(self):
        buffer = create_excel([_make_bid()])
        wb = load_workbook(buffer)
        assert wb is not None
        wb.close()

    def test_has_two_sheets(self):
        buffer = create_excel([_make_bid()])
        wb = load_workbook(buffer)
        assert len(wb.sheetnames) == 2
        assert "Licitações Uniformes" in wb.sheetnames
        assert "Metadata" in wb.sheetnames
        wb.close()

    def test_header_count(self):
        buffer = create_excel([_make_bid()])
        wb = load_workbook(buffer)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert len([h for h in headers if h]) == 11
        wb.close()

    def test_header_names(self):
        expected = ["Código PNCP", "Objeto", "Órgão", "UF", "Município",
                     "Valor Estimado", "Modalidade", "Publicação", "Início",
                     "Situação", "Link"]
        buffer = create_excel([_make_bid()])
        wb = load_workbook(buffer)
        ws = wb.active
        actual = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        assert actual == expected
        wb.close()

    def test_header_green_fill(self):
        buffer = create_excel([_make_bid()])
        wb = load_workbook(buffer)
        ws = wb.active
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            # openpyxl returns colors with or without FF prefix
            assert cell.fill.start_color.rgb in ("002E7D32", "FF2E7D32")
        wb.close()

    def test_frozen_panes(self):
        buffer = create_excel([_make_bid()])
        wb = load_workbook(buffer)
        ws = wb.active
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_data_rows_correct_count(self):
        bids = [_make_bid(codigoCompra=f"code-{i}") for i in range(5)]
        buffer = create_excel(bids)
        wb = load_workbook(buffer)
        ws = wb.active
        # 1 header + 5 data + 1 total = 7 rows
        assert ws.max_row == 7
        wb.close()

    def test_currency_format(self):
        buffer = create_excel([_make_bid(valorTotalEstimado=100000.50)])
        wb = load_workbook(buffer)
        ws = wb.active
        valor_cell = ws.cell(row=2, column=6)
        assert valor_cell.value == 100000.50
        assert "R$" in (valor_cell.number_format or "")
        wb.close()

    def test_total_row_sum_formula(self):
        bids = [_make_bid(valorTotalEstimado=100.0) for _ in range(3)]
        buffer = create_excel(bids)
        wb = load_workbook(buffer)
        ws = wb.active
        total_row = len(bids) + 2  # header + data
        total_cell = ws.cell(row=total_row, column=6)
        assert "SUM" in str(total_cell.value)
        wb.close()

    def test_hyperlink_column(self):
        buffer = create_excel([_make_bid(linkSistemaOrigem="https://example.com/edital")])
        wb = load_workbook(buffer)
        ws = wb.active
        link_cell = ws.cell(row=2, column=11)
        assert link_cell.value == "Abrir"
        assert link_cell.hyperlink is not None
        wb.close()

    def test_empty_list(self):
        buffer = create_excel([])
        wb = load_workbook(buffer)
        ws = wb.active
        assert ws.max_row == 1  # only header
        wb.close()

    def test_invalid_input_raises(self):
        with pytest.raises(ValueError):
            create_excel("not a list")

    def test_large_dataset_500_rows(self):
        bids = [_make_bid(codigoCompra=f"bid-{i}", valorTotalEstimado=float(i * 1000)) for i in range(500)]
        buffer = create_excel(bids)
        wb = load_workbook(buffer)
        ws = wb.active
        assert ws.max_row == 502  # header + 500 + total
        wb.close()

    def test_metadata_sheet(self):
        bids = [_make_bid(valorTotalEstimado=100.0) for _ in range(3)]
        buffer = create_excel(bids)
        wb = load_workbook(buffer)
        ws_meta = wb["Metadata"]
        assert ws_meta["A1"].value == "Gerado em:"
        assert ws_meta["A2"].value == "Total de licitações:"
        assert ws_meta["B2"].value == 3
        assert ws_meta["A3"].value == "Valor total estimado:"
        assert ws_meta["B3"].value == 300.0
        wb.close()

    def test_none_values_handled(self):
        bid = _make_bid(
            objetoCompra=None,
            nomeOrgao=None,
            municipio=None,
            valorTotalEstimado=None,
            dataPublicacaoPncp=None,
            dataAberturaProposta=None,
        )
        buffer = create_excel([bid])
        wb = load_workbook(buffer)
        ws = wb.active
        # sanitize_for_excel returns "", but openpyxl stores empty string as None
        assert ws.cell(row=2, column=2).value in ("", None)
        assert ws.cell(row=2, column=3).value in ("", None)
        assert ws.cell(row=2, column=6).value is None  # None valor
        wb.close()

    def test_pncp_link_construction(self):
        """Test fallback PNCP link construction from numeroControlePNCP."""
        bid = _make_bid(
            linkSistemaOrigem=None,
            linkProcessoEletronico=None,
            numeroControlePNCP="67366310000103-1-000189/2025",
        )
        buffer = create_excel([bid])
        wb = load_workbook(buffer)
        ws = wb.active
        link_cell = ws.cell(row=2, column=11)
        assert link_cell.hyperlink is not None
        assert "67366310000103" in str(link_cell.hyperlink.target)
        assert "2025" in str(link_cell.hyperlink.target)
        wb.close()


class TestSanitizeForExcel:
    """Test illegal character sanitization."""

    def test_removes_control_chars(self):
        assert sanitize_for_excel("ABC\x13DEF") == "ABC DEF"

    def test_preserves_normal_text(self):
        assert sanitize_for_excel("Normal text") == "Normal text"

    def test_handles_none(self):
        assert sanitize_for_excel(None) == ""

    def test_handles_non_string(self):
        assert sanitize_for_excel(12345) == "12345"

    def test_multiple_control_chars(self):
        result = sanitize_for_excel("\x00\x01\x02\x03hello\x0b\x0c\x0e")
        assert "hello" in result
        assert "\x00" not in result


class TestParseDatetime:
    """Test datetime parsing from PNCP formats."""

    def test_iso_with_z(self):
        result = parse_datetime("2024-01-25T10:30:00Z")
        assert result == datetime(2024, 1, 25, 10, 30, 0)
        assert result.tzinfo is None  # Must be naive for Excel

    def test_iso_without_timezone(self):
        result = parse_datetime("2024-01-25T10:30:00")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_date_only(self):
        result = parse_datetime("2024-01-25")
        assert result == datetime(2024, 1, 25, 0, 0, 0)

    def test_none_returns_none(self):
        assert parse_datetime(None) is None

    def test_invalid_string_returns_none(self):
        assert parse_datetime("not-a-date") is None

    def test_empty_string_returns_none(self):
        assert parse_datetime("") is None
